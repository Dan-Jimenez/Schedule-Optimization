from pulp import LpProblem, LpMinimize, LpVariable, lpSum, value
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Leer el archivo Excel
df = pd.read_excel("Entregable 3 Horarios Pulp.xlsx")

# Obtener datos únicos de profesores, materias, horarios y salones
profesores = df['Profesores'].unique()
materias = df['Materias'].unique()
horarios = df['Horarios'].unique()
salones = df['Salones'].unique()
disp_materias = df['Materias que puede dar cada profesor']
disp_horario = df['Disponibilidad de horario de cada profesor']

# Crear el problema
prob = LpProblem("Programacion de Horarios", LpMinimize)

# Crear las variables
x = LpVariable.dicts("Asignacion", 
                      [(prof, mat, sal, hor) for prof in profesores 
                                            for mat in materias 
                                            for sal in salones 
                                            for hor in horarios], 
                      cat='Binary')

# Función objetivo: minimizar el costo total
prob += lpSum(x[(prof, mat, sal, hor)] * (
        df.loc[df['Profesores'] == prof, 'Costos Profesores'].iloc[0] +
        df.loc[df['Materias'] == mat, 'Costos Materias'].iloc[0] +
        df.loc[df['Horarios'] == hor, 'Costos Horarios'].iloc[0] +
        df.loc[df['Salones'] == sal, 'Costos Salones'].iloc[0])
        for prof in profesores for mat in materias for hor in horarios for sal in salones)

# Restricciones
# 1. Cada materia solo puede ser enseñada una vez
for mat in materias:
    prob += lpSum(x[(prof, mat, sal, hor)] for prof in profesores for sal in salones for hor in horarios) == 1

# 2. Cada horario solo puede ser utilizado una vez en algún salón
for hor in horarios:
    prob += lpSum(x[(prof, mat, sal, hor)] for prof in profesores for mat in materias for sal in salones) == 1

# 3. Restricciones específicas de los profesores en cuanto a materias y horarios
for prof, disp_mats, disp_hors in zip(profesores, disp_materias, disp_horario):
    disp_mats = [char for char in disp_mats]
    disp_hors = [int(char) for char in str(disp_hors)]
    for mat in materias:
        if mat not in disp_mats:
            for hor in horarios:
                prob += lpSum(x[(prof, mat, sal, hor)] for sal in salones) == 0
        else:
            for hor in [h for h in horarios if h not in disp_hors]:
                prob += lpSum(x[(prof, mat, sal, hor)] for sal in salones) == 0

# Resolver el problema
prob.solve()
costo = value(prob.objective)
# Mostrar resultados
print("Costo total:", costo)
print("\nHorario:")
for sal in salones:
    for hor in horarios:
        for prof in profesores:
            for mat in materias:
                if value(x[(prof, mat, sal, hor)]) == 1:
                    print(f"Salón {sal}, Horario {hor}, Profesor {prof}, Materia {mat}")

# Nombre del archivo Excel
excel_file = "horarios_asignados.xlsx"

try:
    # Cargar el archivo de Excel en caso de existir
    wb = load_workbook(excel_file)
    ws = wb.active
    # Eliminar el contenido existente en el archivo Excel
    ws.delete_rows(ws.min_row, ws.max_row)
except FileNotFoundError:
    # Crear un nuevo archivo de Excel en caso de que no exista uno
    wb = Workbook()
    ws = wb.active
    ws.title = "Horarios Asignados"

# Añadir los nombres de las columnas
ws.append(["Costo", costo])
ws.append(["Profesor", "Materia", "Horario", "Salón"])
for row in ws.iter_rows(min_row=1, max_row=2):
    for cell in row:
        cell.font = Font(size=14, bold=True)
# Escribir los resultados en el archivo Excel
for sal in salones:
    for hor in horarios:
        for prof in profesores:
            for mat in materias:
                if value(x[(prof, mat, sal, hor)]) == 1:
                    ws.append([prof, mat, hor, sal])
for row in ws.iter_rows(min_row=3, max_row=12):
    for cell in row:
        cell.font = Font(size=14)

# Guardar el archivo Excel
wb.save(excel_file)

print("Los horarios asignados se han guardado en 'horarios_asignados.xlsx'")